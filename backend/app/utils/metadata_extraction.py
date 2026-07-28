import logging
from datetime import date
from datetime import datetime
from typing import Any

logger = logging.getLogger("app.utils.metadata_extraction")


# ==========================================================
# JSON-safety
# ==========================================================
#
# Root cause of the Reverse Image 502: this dict is persisted straight
# into a plain SQLAlchemy `JSON` column (InvestigationResult.data and
# ImageFingerprint.extracted_metadata), which serializes with the
# stdlib `json` module. Real-world camera/phone EXIF routinely contains
# types `json.dumps` cannot handle on its own - most commonly
# `PIL.TiffImagePlugin.IFDRational` (ExposureTime, FNumber, FocalLength,
# and GPSLatitude/GPSLongitude/GPSAltitude are all rationals) and raw
# `bytes` for a handful of GPS sub-tags (e.g. GPSAltitudeRef,
# GPSProcessingMethod). The unit tests only ever exercised
# PIL-generated PNGs with zero EXIF, so this never surfaced there - it
# only appears once a real photograph is uploaded, which matches the
# reported "worked, then flipped to 502" behavior. `db.commit()` then
# raises a `TypeError` from deep inside the SQLAlchemy/sqlite3 JSON
# encoder, which the endpoint's blanket `except Exception` turns into
# an opaque 502.
#
# `_json_safe` recursively coerces a value into something `json.dumps`
# is guaranteed to accept, so metadata extraction can never again take
# down persistence for a valid image just because of an unusual tag
# type.


def _json_safe(value: Any) -> Any:

    if value is None or isinstance(value, (str, bool, int, float)):
        return value

    if isinstance(value, bytes):

        try:
            return value.decode("utf-8", errors="replace")
        except Exception:
            return repr(value)

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, dict):
        return {str(_json_safe(k)): _json_safe(v) for k, v in value.items()}

    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]

    # Covers PIL's IFDRational and any other numerator/denominator-style
    # rational (e.g. fractions.Fraction): render as a plain float rather
    # than failing to serialize or silently dropping precision info.
    if hasattr(value, "numerator") and hasattr(value, "denominator"):

        try:
            return float(value)
        except Exception:
            return str(value)

    try:
        float(value)
        return float(value)
    except (TypeError, ValueError):
        pass

    # Last resort: never let an unrecognized EXIF/metadata value type
    # crash the investigation. A readable string beats a lost record.
    return str(value)



# ==========================================================
# Image EXIF
# ==========================================================

def extract_image_metadata(path: str) -> dict[str, Any]:

    from PIL import Image
    from PIL import ExifTags

    result: dict[str, Any] = {}

    try:
        with Image.open(path) as image:

            result["format"] = image.format
            result["mode"] = image.mode
            result["width"] = image.width
            result["height"] = image.height

            exif_data = image.getexif()

            if exif_data:

                readable: dict[str, Any] = {}

                for tag_id, value in exif_data.items():

                    tag_name = ExifTags.TAGS.get(tag_id, str(tag_id))

                    if isinstance(value, bytes):

                        try:
                            value = value.decode(errors="replace")
                        except Exception:
                            value = repr(value)

                    readable[tag_name] = value

                result["exif"] = readable

                gps_info = exif_data.get_ifd(ExifTags.IFD.GPSInfo) if hasattr(
                    exif_data, "get_ifd"
                ) else None

                if gps_info:
                    result["gps"] = {
                        ExifTags.GPSTAGS.get(k, str(k)): v
                        for k, v in gps_info.items()
                    }

    except Exception as error:
        logger.warning("Image metadata extraction failed: %s", error)
        result["error"] = str(error)

    return result


# ==========================================================
# PDF
# ==========================================================

def extract_pdf_metadata(path: str) -> dict[str, Any]:

    from pypdf import PdfReader

    result: dict[str, Any] = {}

    try:
        reader = PdfReader(path)

        info = reader.metadata or {}

        result["page_count"] = len(reader.pages)
        result["title"] = info.get("/Title")
        result["author"] = info.get("/Author")
        result["subject"] = info.get("/Subject")
        result["creator"] = info.get("/Creator")
        result["producer"] = info.get("/Producer")
        result["creation_date"] = _stringify(info.get("/CreationDate"))
        result["modification_date"] = _stringify(info.get("/ModDate"))
        result["is_encrypted"] = reader.is_encrypted

    except Exception as error:
        logger.warning("PDF metadata extraction failed: %s", error)
        result["error"] = str(error)

    return result


# ==========================================================
# DOCX
# ==========================================================

def extract_docx_metadata(path: str) -> dict[str, Any]:

    from docx import Document

    result: dict[str, Any] = {}

    try:
        document = Document(path)
        props = document.core_properties

        result["title"] = props.title
        result["author"] = props.author
        result["subject"] = props.subject
        result["keywords"] = props.keywords
        result["comments"] = props.comments
        result["last_modified_by"] = props.last_modified_by
        result["revision"] = props.revision
        result["created"] = _stringify(props.created)
        result["modified"] = _stringify(props.modified)
        result["paragraph_count"] = len(document.paragraphs)

    except Exception as error:
        logger.warning("DOCX metadata extraction failed: %s", error)
        result["error"] = str(error)

    return result


# ==========================================================
# PPTX
# ==========================================================

def extract_pptx_metadata(path: str) -> dict[str, Any]:

    from pptx import Presentation

    result: dict[str, Any] = {}

    try:
        presentation = Presentation(path)
        props = presentation.core_properties

        result["title"] = props.title
        result["author"] = props.author
        result["subject"] = props.subject
        result["keywords"] = props.keywords
        result["comments"] = props.comments
        result["last_modified_by"] = props.last_modified_by
        result["created"] = _stringify(props.created)
        result["modified"] = _stringify(props.modified)
        result["slide_count"] = len(presentation.slides)

    except Exception as error:
        logger.warning("PPTX metadata extraction failed: %s", error)
        result["error"] = str(error)

    return result


# ==========================================================
# XLSX
# ==========================================================

def extract_xlsx_metadata(path: str) -> dict[str, Any]:

    from openpyxl import load_workbook

    result: dict[str, Any] = {}

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        props = workbook.properties

        result["title"] = props.title
        result["creator"] = props.creator
        result["subject"] = props.subject
        result["keywords"] = props.keywords
        result["last_modified_by"] = props.lastModifiedBy
        result["created"] = _stringify(props.created)
        result["modified"] = _stringify(props.modified)
        result["sheet_names"] = workbook.sheetnames

        workbook.close()

    except Exception as error:
        logger.warning("XLSX metadata extraction failed: %s", error)
        result["error"] = str(error)

    return result


# ==========================================================
# Dispatcher
# ==========================================================

_MIME_DISPATCH = {
    "image/jpeg": extract_image_metadata,
    "image/png": extract_image_metadata,
    "image/gif": extract_image_metadata,
    "image/tiff": extract_image_metadata,
    "image/bmp": extract_image_metadata,
    "image/webp": extract_image_metadata,
    "application/pdf": extract_pdf_metadata,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": extract_docx_metadata,
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": extract_pptx_metadata,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": extract_xlsx_metadata,
}

_EXTENSION_FALLBACK = {
    ".jpg": extract_image_metadata,
    ".jpeg": extract_image_metadata,
    ".png": extract_image_metadata,
    ".gif": extract_image_metadata,
    ".bmp": extract_image_metadata,
    ".tiff": extract_image_metadata,
    ".webp": extract_image_metadata,
    ".pdf": extract_pdf_metadata,
    ".docx": extract_docx_metadata,
    ".pptx": extract_pptx_metadata,
    ".xlsx": extract_xlsx_metadata,
}


def extract_metadata(
    *,
    path: str,
    detected_mime_type: str | None,
    declared_extension: str | None,
) -> dict[str, Any]:
    """
    Picks the right extractor by sniffed MIME type first (never trusts
    the client-declared Content-Type), falling back to the file
    extension only if the MIME type wasn't recognized. Unsupported file
    types return an empty, valid result rather than an error - metadata
    extraction being unavailable for a type is not a failure of the
    investigation.
    """

    handler = _MIME_DISPATCH.get(detected_mime_type or "")

    if handler is None and declared_extension:
        handler = _EXTENSION_FALLBACK.get(declared_extension)

    if handler is None:
        return {"supported": False}

    metadata = handler(path)
    metadata["supported"] = True

    # Every extractor above pulls values straight from third-party
    # libraries (Pillow EXIF, pypdf info dict, python-docx/pptx core
    # properties, openpyxl properties) whose types aren't guaranteed to
    # be JSON-serializable. Sanitize once here, in the single place all
    # callers (ReverseImageIntelligenceService, File Intelligence, etc.)
    # go through, rather than trusting every extractor to do it
    # individually.
    return _json_safe(metadata)


def _stringify(value: Any) -> str | None:

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    return str(value)
